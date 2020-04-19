import os
import shutil
import openpyxl


def copy_file_and_rename(file_name : str):
    if not os.path.exists(file_name):
        print(file_name, '不存在')
        return
    
    # src_name = os.path.basename(file_name)
    # print(src_name)

    dir_name = os.path.dirname(file_name)
    image_names = get_image_names(dir_name)

    for name in image_names:
        # print(name)
        shutil.copy(file_name, os.path.join(dir_name, name + '.jpg'))

    pass


def get_image_names(file_path : str):
    print(file_path)
    
    wb = openpyxl.load_workbook(file_path + '\\' + '人员信息导入(模板).xlsx')
    ws = wb[wb.sheetnames[0]]

    rows = ws.max_row
    cols = ws.max_column
    # print(rows)
    # print(cols)
    # sheet_display(ws)

    image_names = []
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            if col == 3 and (row >= 2 and row <= rows - 2):
                # print(ws.cell(row, col).value)
                image_names.append(str(ws.cell(row, col).value))

    return image_names
    pass


def sheet_display(ws: openpyxl.worksheet.worksheet.Worksheet):
    rows = ws.max_row
    cols = ws.max_column
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            print(ws.cell(row, col).value, end='\t')
        print()


if __name__ == '__main__':
    # copy_file_and_rename(input('请输入文件名称: '))
    # C:\Users\DLAX\Desktop\batch import\image.jpg
    copy_file_and_rename(r'C:\Users\DLAX\Desktop\batch import\image.jpg')
    pass

