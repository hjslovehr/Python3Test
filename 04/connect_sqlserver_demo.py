
# import pymssql
# DeprecationWarning: Using or importing the ABCs from 'collections' 
# instead of from 'collections.abc' is deprecated, and in 3.8 it will 
# stop working
import sys
sys.path.append('.\\common')
import sqlhelper
import openpyxl


server = '123.56.96.237'
database = 'DFMSDB'
user = 'sa'
password = 'Simp2014'


def query_test():
    con = sqlhelper.get_connection(server, database, user, password)
    # res = sqlhelper.get_all(con, 'SELECT PERSONNAME, PERSONID, PERSONCODE FROM BASE_PERSON')
    res = sqlhelper.get_all(con, 'SELECT * FROM BASE_PERSON ')
    con.close()

    cols = len(res[0])
    for item in res:
        for i in range(cols):
            print(item[i], end='\t')
        print()

    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = "persons"

    for row in range(1, len(res) + 1):
        for col in range(1, cols + 1):
            ws.cell(row, col).value = str(res[row - 1][col - 1])

    wb.save(r'C:\Users\DLAX\Desktop\person.xlsx')
    wb.close()
    pass



if __name__ == '__main__':
    query_test()
    pass



