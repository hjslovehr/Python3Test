import openpyxl


def read_test(excel_name: str):
    wb = openpyxl.load_workbook(excel_name)
    print(wb.sheetnames)
    ws = wb[wb.sheetnames[0]]

    sheet_display(ws)

    pass


def sheet_display(ws: openpyxl.worksheet.worksheet.Worksheet):
    rows = ws.max_row
    cols = ws.max_column
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            print(ws.cell(row, col).value, end='\t')
        print()


if __name__ == '__main__':
    read_test(r'C:\Users\DLAX\Desktop\123.xlsx')
    pass


