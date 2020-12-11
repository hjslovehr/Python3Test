import openpyxl as excel

def write_icno():
    iccard_data_file = r'C:\Users\Lenovo\Desktop\新建文件夹\门禁人员信息.xlsx'
    fastgate_data_file = r'C:\Users\Lenovo\Desktop\新建文件夹\personxxx.xlsx'

    wb = excel.load_workbook(fastgate_data_file)
    ws = wb[wb.sheetnames[0]]

    card_nos = get_cardnos(iccard_data_file)

    keys = card_nos.keys()

    # 调试打印速通门人和卡信息
    # for key in keys:
    #     print(key, '==>', card_nos[key])
    
    print('获取门禁卡信息：', len(card_nos), '条')

    count = 0
    # 修改数据
    for row in range(1 + 1, ws.max_row + 1):
        if str(ws.cell(row, 1).value) in keys:
            # print(str(ws.cell(row, 1).value), str(ws.cell(row, 2).value))
            count += 1
            cardno = ''
            if (None == card_nos[str(ws.cell(row, 1).value)] 
                or 'None' == card_nos[str(ws.cell(row, 1).value)]):
                cardno = ''
            else:
                cardno = card_nos[str(ws.cell(row, 1).value)]
            ws.cell(row, 7).value = cardno
            pass
    
    print('匹配到速通门人员信息：', count , '条')
            
    # 保存 Excel 文件
    wb.save(fastgate_data_file)

    pass


def get_cardnos(excel_file:str):
    wb = excel.load_workbook(excel_file)
    ws = wb[wb.sheetnames[0]]
    card_nos = {}
    for row in range(1 + 1, ws.max_row + 1):
        card_nos[str(ws.cell(row, 3).value)] = str(ws.cell(row, 12).value)
    
    return card_nos
    pass


if __name__ == "__main__":
    write_icno()
    pass

