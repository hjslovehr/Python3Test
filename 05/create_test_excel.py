import openpyxl as excel


def create_test_data(max_sheet_count:int, max_row_count:int):
    print('最大 sheet 数目:', max_sheet_count)
    print('最大记录行数:', max_row_count)
    
    wb = excel.Workbook()
    for i in range(1 + 1, max_sheet_count + 1):
        wb.create_sheet(f'Sheet{i}')
    
    for name in wb.sheetnames:
        ws = wb[name]
        for row in range(1, max_row_count + 1):
            for col in range(1, 8 + 1):
                ws.cell(row, col).value = 'ABCDEFGH'
            pass
        pass

    wb.save(r'C:\Users\DLAX\Desktop\新建文件夹\temp.xlsx')
    pass


if __name__ == "__main__":
    create_test_data(int(input('请输入最大 sheet 数目: ')), 
                     int(input('请输入最大记录行数: ')))
    pass


