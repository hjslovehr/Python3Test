import openpyxl


def person_compare():
    '''人员对比'''
    # result_file = r'C:\Users\Lenovo\Desktop\新建文件夹\result.xlsx'
    # card_data_file = r'C:\Users\Lenovo\Desktop\新建文件夹\员工门禁卡数据.xlsx'
    # result_excel = r'C:\Users\Lenovo\Desktop\新建文件夹\人员对比结果.xlsx'

    result_file = r'C:\Users\DLAX\Desktop\新建文件夹\result.xlsx'
    card_data_file = r'C:\Users\DLAX\Desktop\新建文件夹\员工门禁卡数据.xlsx'
    result_excel = r'C:\Users\DLAX\Desktop\新建文件夹\人员对比结果.xlsx'

    print('获取速通门人员数据...')
    # 取 result.xlsx 中的数据
    fastgate_persons = get_fastgate_persons(result_file)

    print('获取门禁人员编号...')
    # 取员工门禁卡数据中的人员编号
    card_data_persons = get_carddata_persons(card_data_file)
    
    # debug print
    # for key in fastgate_persons.keys():
    #     print(key, fastgate_persons[key])

    # debug print
    # for item in card_data_persons:
    #     print(item)

    print('数据对比...')
    result_data = data_compare(fastgate_persons, card_data_persons)

    # debug print
    # for key in result_data.keys():
    #     print(key, result_data[key])

    # debug print
    # for key in result_data.keys():
    #     print(key, '==>', 
    #           result_data[key]['PersonCode'], 
    #           result_data[key]['DptName'], 
    #           result_data[key]['DptCode'], 
    #           result_data[key]['PersonName'], 
    #           result_data[key]['Gender'], 
    #           result_data[key]['ICCardNo'], 
    #           result_data[key]['IDNO'])
    #     pass 

    print('速通门人员:', len(fastgate_persons), '个')
    print('门禁人员:', len(card_data_persons), '个')
    print('速通门人员', len(result_data), '个人不在门禁系统中')

    # 保存文件
    save_result_to_excel(result_excel, result_data)
    print(f"保存{result_excel}")

    pass


def get_fastgate_persons(excel_file:str):
    '''获取速通门人员'''
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[wb.sheetnames[0]]
    rows = ws.max_row
    result_persons = {}
    for row in range(1 + 1, rows + 1):
        result_persons[str(ws.cell(row, 1).value)] = {}
        temp = result_persons[str(ws.cell(row, 1).value)]
        temp['PersonCode'] = str(ws.cell(row, 1).value)
        temp['DptName'] = str(ws.cell(row, 2).value)
        temp['DptCode'] = str(ws.cell(row, 3).value)
        temp['PersonName'] = str(ws.cell(row, 4).value)
        temp['Gender'] = str(ws.cell(row, 5).value)
        temp['ICCardNo'] = str(ws.cell(row, 6).value)
        temp['IDNO'] = str(ws.cell(row, 7).value)
        pass

    return result_persons


def get_carddata_persons(excel_file:str):
    '''获取门禁人员'''
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[wb.sheetnames[3]]
    rows = ws.max_row
    card_data_persons = []
    for row in range(1, rows + 1):
        card_data_persons.append(ws.cell(row, 2).value)
    
    return card_data_persons


def data_compare(fastgate_persons:dict, card_data_persons:list):
    '''数据对比'''
    result = {}
    for key in fastgate_persons.keys():
        if key not in card_data_persons:
            result[key] = fastgate_persons[key]
    
    return result


def save_result_to_excel(excel_file:str, data:dict):
    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = "结果"

    row = 1
    ws.cell(row, 1).value = 'PersonCode'
    ws.cell(row, 2).value = 'DptName'
    ws.cell(row, 3).value = 'DptCode'
    ws.cell(row, 4).value = 'PersonName'
    ws.cell(row, 5).value = 'Gender'
    ws.cell(row, 6).value = 'ICCardNo'
    ws.cell(row, 7).value = 'IDNO'

    row += 1
    for key in data.keys():
        ws.cell(row, 1).value = str(data[key]['PersonCode'])
        ws.cell(row, 2).value = (str(data[key]['DptName']) if 'None' != str(data[key]['DptName']) else '')
        ws.cell(row, 3).value = (str(data[key]['DptCode']) if 'None' != str(data[key]['DptCode']) else '')
        ws.cell(row, 4).value = (str(data[key]['PersonName']) if 'None' != str(data[key]['PersonName']) else '')
        ws.cell(row, 5).value = (str(data[key]['Gender']) if 'None' != str(data[key]['Gender']) else '')
        ws.cell(row, 6).value = (str(data[key]['ICCardNo']) if 'None' != str(data[key]['ICCardNo']) else '')
        ws.cell(row, 7).value = (str(data[key]['IDNO']) if 'None' != str(data[key]['IDNO']) else '')
        row += 1
        pass

    wb.save(excel_file)
    pass


if __name__ == "__main__":
    person_compare()
    pass
