import openpyxl

column_names = [
    '部门', '人员编号', '姓名', '性别', '职位', '卡号'
]

def person_compare():
    '''人员对比'''
    # result_file = r'C:\Users\Lenovo\Desktop\新建文件夹\result.xlsx'
    # card_data_file = r'C:\Users\Lenovo\Desktop\新建文件夹\员工门禁卡数据.xlsx'
    # result_excel = r'C:\Users\Lenovo\Desktop\新建文件夹\人员对比结果.xlsx'

    result_file = r'C:\Users\DLAX\Desktop\新建文件夹\result.xlsx'
    card_data_file = r'C:\Users\DLAX\Desktop\新建文件夹\员工门禁卡数据.xlsx'
    result_excel = r'C:\Users\DLAX\Desktop\新建文件夹\人员对比结果.xlsx'

    print('获取速通门人员数据...')
    # 取 result.xlsx 中的数据
    fastgate_persons = get_fastgate_persons(result_file)

    print('获取门禁人员编号...')
    # 取员工门禁卡数据中的人员编号
    card_data_persons = get_carddata_persons(card_data_file)
    
    # debug print
    # for item in card_data_persons:
    #     print(item)
    
    # debug print
    # for key in card_data_persons.keys():
    #     print(key, card_data_persons[key])

    print('数据对比...')
    result_data = data_compare(fastgate_persons, card_data_persons)

    # debug print
    # for key in result_data.keys():
    #     print(key, result_data[key])

    # debug print
    print('打印差异数据...')
    for key in result_data.keys():
        # print(key, result_data[key])
        print(key, '==>', 
              result_data[key][column_names[0]], 
              result_data[key][column_names[1]], 
              result_data[key][column_names[2]], 
              result_data[key][column_names[3]], 
              result_data[key][column_names[4]], 
              result_data[key][column_names[5]])
        pass 

    print('速通门人员:', len(fastgate_persons), '个')
    print('门禁人员:', len(card_data_persons), '个')
    print('速通门人员', len(result_data), '个人不在门禁系统中')

    # 保存文件
    save_result_to_excel(result_excel, result_data)
    print(f"保存{result_excel}")

    pass


def get_fastgate_persons(excel_file:str):
    '''获取速通门人员'''
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[wb.sheetnames[0]]
    rows = ws.max_row
    fastgate_persons = []
    for row in range(1 + 1, rows + 1):
        fastgate_persons.append(str(ws.cell(row, 1).value))
        pass

    return fastgate_persons


def get_carddata_persons(excel_file:str):
    '''获取门禁人员'''
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[wb.sheetnames[3]]
    rows = ws.max_row
    card_data_persons = {}
    for row in range(1, rows + 1):
        temp = {}
        temp[column_names[0]] = ws.cell(row, 1).value
        temp[column_names[1]] = ws.cell(row, 2).value
        temp[column_names[2]] = ws.cell(row, 3).value
        temp[column_names[3]] = ws.cell(row, 4).value
        temp[column_names[4]] = ws.cell(row, 5).value
        temp[column_names[5]] = ws.cell(row, 9).value

        card_data_persons[str(ws.cell(row, 2).value)] = temp
        pass
    
    return card_data_persons


def data_compare(fastgate_persons:list, card_data_persons:dict):
    '''数据对比'''
    result = {}
    for key in card_data_persons.keys():
        if key not in fastgate_persons:
            result[key] = card_data_persons[key]
    
    return result


def save_result_to_excel(excel_file:str, data:dict):
    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[0]]
    ws.title = "结果"

    row = 1
    col_index = 1
    for item in column_names:
        ws.cell(row, col_index).value = item
        col_index += 1

    row += 1
    for key in data.keys():
        for index in range(0, len(column_names)):
            ws.cell(row, index + 1).value = (str(data[key][column_names[index]]) if None != data[key][column_names[1]] else '')
            pass
        row += 1
        pass

    wb.save(excel_file)
    pass


if __name__ == "__main__":
    person_compare()
    pass
